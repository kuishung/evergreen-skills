#!/usr/bin/env python3
"""
compile_cfp.py
==============

Skill: cfp-entry  (version 0.7.0)

CFP voucher-redemption -> AutoCount Sales Invoice import compiler.

What it does
------------
1. Reads one or more daily CFP / gvLedger source files. Two formats are
   supported (per file extension):
     * PDF  -- TK CFP REPORT, gvLedger BS, gvLedger BL (one file per
               station, the original v0.1+ format).
     * XLSX -- the unified gvLedger export (v0.7.0+) -- ONE file covering
               all three stations, with a Station column on each row.
2. Parses every voucher-redemption line.
3. Looks up each company against customer_codes.csv (extracted from the
   AutoCount Debtor Listing) to attach the 300-XXXX account code.
4. Groups all redemptions by (account_code, gas_type) for the day so each
   customer appears in AutoCount as one Petrol line and/or one Diesel
   line per day -- never one line per voucher.
5. Opens the user's AutoCount "Default Sales.xlsx" template, copies its
   header layout, and writes the consolidated rows into a new xlsx
   ready for AutoCount's Sales Invoice import.

Usage
-----
    python compile_cfp.py \
        --reports "D:/CLAUDE/Skill/CFP Entry/20260501-CFP REPORT-TK.pdf" \
                  "D:/CLAUDE/Skill/CFP Entry/gvLedger - 2026-05-01T143652.120.pdf" \
                  "D:/CLAUDE/Skill/CFP Entry/gvLedger - BL.pdf" \
        --date 2026-05-01 \
        --template "D:/CLAUDE/Skill/CFP Entry/Autocount Import Template/Default Sales.xlsx" \
        --customers "D:/CLAUDE/Skill/CFP Entry/cfp-entry/customer_codes.csv" \
        --out "D:/CLAUDE/Skill/CFP Entry/output/SalesImport_2026-05-01.xlsx" \
        --petrol-code "" --diesel-code ""

If --date is omitted, the script infers it from the first redemption row.

Customer-name matching
----------------------
The CFP report and AutoCount sometimes spell names slightly differently
(extra spaces, "(SABAH)" added, trailing "- (BL-V14)" etc.). The
matcher therefore:
  * uppercases & collapses whitespace on both sides
  * strips trailing parenthetical "- (BL-V14)" / "(BS-V1)" markers
  * tries exact match first, then a fuzzy contains/Jaccard fallback
  * any unmatched company is written into a separate "Unmapped" sheet
    so you can fix the master list and re-run

Dependencies
------------
    pip install -r requirements.txt
or:
    pip install pdfplumber openpyxl rapidfuzz

(rapidfuzz is optional; if unavailable, a built-in fallback is used.
 pdfplumber is only needed when a PDF source is passed in -- xlsx-only
 runs do not require it, but the import is unconditional today.)

Author : Evergreen back-office automation
Skill  : cfp-entry
See    : SKILL.md and CHANGELOG.md for version history.
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

# pdfplumber is imported lazily inside parse_pdf so an xlsx-only run
# (the v0.7.0+ default) doesn't require it. openpyxl is always required
# (input parser for xlsx sources + writer for the AutoCount import).
pdfplumber = None  # set on first PDF parse; left None for xlsx-only runs

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    print("ERROR: openpyxl is required. Run: pip install openpyxl", file=sys.stderr)
    raise

try:
    from rapidfuzz import fuzz as _rfz_fuzz
    def _ratio(a: str, b: str) -> float:
        return _rfz_fuzz.token_set_ratio(a, b) / 100.0
except ImportError:  # pragma: no cover
    def _ratio(a: str, b: str) -> float:
        # crude Jaccard on tokens as a fallback
        sa, sb = set(a.split()), set(b.split())
        if not sa or not sb:
            return 0.0
        return len(sa & sb) / len(sa | sb)


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class Redemption:
    company_raw: str
    gas_type: str          # "Petrol" / "Diesel"
    amount: float          # absolute, positive
    litre: Optional[float] # see notes; may be filled in post-parse via RefPrice
    vehicle: str
    station: str           # "Buraqoil Tg Kapor" / "...Berkat Setia" / "...Bubul Lama"
    voucher: str
    redeem_dt: datetime
    receipt: str
    # v0.7.0: True when `litre` was derived from RefPrice rather than read
    # from the source. Three cases set this True:
    #   * TK / BS PDF rows (the source format has no litre column at all);
    #   * xlsx rows with a blank or zero Litre cell (per-row gap);
    #   * any future source where per-row litres are missing.
    # Set in `derive_missing_litres()` which runs once after all sources
    # are parsed. Affects: unit_price_per_row (returns None for estimated
    # rows so Check C exempts them); FurtherDescription voucher line
    # (shows "RM X.XX" instead of "X.XXL" -- never report a derived
    # litre as if it were source truth); Description suffix ("est" / "n est").
    litre_estimated: bool = False

    @property
    def date(self) -> str:
        return self.redeem_dt.strftime("%Y-%m-%d")

    @property
    def unit_price_per_row(self) -> Optional[float]:
        """v0.6.3: per-source-row unit price = amount / litre, rounded to 2 dp.
        Returns None when the source row has no usable litre value -- either
        because the source format omits litres entirely (TK/BS PDFs), the
        source row has a blank/zero litre (xlsx gap), OR the litre was
        derived from RefPrice (v0.7.0 -- in that case amount/litre would
        just give back the RefPrice and the check would be circular).
        This is the canonical source-of-truth unit price for the new
        UnitPrice rule -- bucket UnitPrice must equal every per-row value
        within the reconciliation tolerance."""
        if self.litre is None or self.litre <= 0 or self.litre_estimated:
            return None
        return round(self.amount / self.litre, 2)


@dataclass
class CustomerRow:
    acc_code: str
    company_name: str
    voucher_code: str
    station: str


# ---------------------------------------------------------------------------
# Per-station fuel item codes  (from Item Listing.pdf, 2026-05-08)
# Spacing is verbatim from AutoCount -- DO NOT normalise.
# ---------------------------------------------------------------------------
STATION_FUEL_TO_ITEM_CODE: Dict[Tuple[str, str], str] = {
    ("TK", "Petrol"): "PETROL - TK",
    ("TK", "Diesel"): "DIESEL -TK",
    ("BS", "Petrol"): "PETROL -BS",
    ("BS", "Diesel"): "DIESEL- BS",
    ("BL", "Petrol"): "PETROL - BL",
    ("BL", "Diesel"): "DIESEL- BL",
}

STATION_NAME_TO_KEY: Dict[str, str] = {
    "Buraqoil Tg Kapor":     "TK",
    "Buraqoil Berkat Setia": "BS",
    "Buraqoil Bubul Lama":   "BL",
}


def load_stock_codes(path: Optional[Path]) -> Dict[Tuple[str, str], str]:
    """Override the built-in mapping with values from stock_codes.csv (if given)."""
    if path is None or not path.exists():
        return dict(STATION_FUEL_TO_ITEM_CODE)
    out: Dict[Tuple[str, str], str] = {}
    with path.open(newline="", encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            stn = (r.get("station") or "").strip().upper()
            gas = (r.get("gas_type") or "").strip().title()
            code = (r.get("item_code") or "").strip()
            if stn and gas and code:
                out[(stn, gas)] = code
    # Fill any gaps from the built-in map
    for k, v in STATION_FUEL_TO_ITEM_CODE.items():
        out.setdefault(k, v)
    return out


def load_project_codes(path: Optional[Path]) -> Dict[str, str]:
    """station -> ProjectNo. Empty values mean 'leave the column blank'."""
    out: Dict[str, str] = {"TK": "", "BS": "", "BL": ""}
    if path is None or not path.exists():
        return out
    with path.open(newline="", encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            stn = (r.get("station") or "").strip().upper()
            code = (r.get("project_code") or "").strip()
            if stn:
                out[stn] = code
    return out


# ---------------------------------------------------------------------------
# Reference prices (v0.6.0): used when the source PDF doesn't carry a litre
# column (TK CFP REPORT, BS gvLedger). Default values inferred from the
# 2026-05-01 BL data and verified to round-trip against the same-day TK/BS
# amounts.
# ---------------------------------------------------------------------------
DEFAULT_REFERENCE_PRICES: Dict[Tuple[str, str], float] = {
    ("TK", "Petrol"): 3.97,
    ("TK", "Diesel"): 2.15,
    ("BS", "Petrol"): 3.97,
    ("BS", "Diesel"): 2.15,
    ("BL", "Petrol"): 3.97,
    ("BL", "Diesel"): 2.15,
}


def load_reference_prices(path: Optional[Path]) -> Dict[Tuple[str, str], float]:
    if path is None or not path.exists():
        return dict(DEFAULT_REFERENCE_PRICES)
    out: Dict[Tuple[str, str], float] = {}
    with path.open(newline="", encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            stn = (r.get("station") or "").strip().upper()
            gas = (r.get("gas_type") or "").strip().title()
            try:
                price = float((r.get("price_per_litre") or "0").strip())
            except ValueError:
                price = 0.0
            if stn and gas and price > 0:
                out[(stn, gas)] = price
    for k, v in DEFAULT_REFERENCE_PRICES.items():
        out.setdefault(k, v)
    return out


# ---------------------------------------------------------------------------
# Customer master loader
# ---------------------------------------------------------------------------

def load_customers(path: Path) -> List[CustomerRow]:
    rows: List[CustomerRow] = []
    with path.open(newline="", encoding="utf-8") as fh:
        for r in csv.DictReader(fh):
            rows.append(CustomerRow(
                acc_code=r["acc_code"].strip(),
                company_name=r["company_name"].strip(),
                voucher_code=(r.get("voucher_code") or "").strip(),
                station=(r.get("station") or "").strip(),
            ))
    return rows


def _norm(s: str) -> str:
    s = s.upper()
    # strip trailing "- (BS-V14)" / "(TK)" / "BS-V14"
    s = re.sub(r"[\(\-]\s*[A-Z]{2}\s*-?\s*V?\s*\d*\s*\)?\s*$", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = s.replace(".", "").replace(",", "")
    s = re.sub(r"\bSDN\.?\s*BHD\.?", "SDN BHD", s)
    s = re.sub(r"\bSDH\b", "SDN", s)  # typo seen in master list
    return s.strip()


class CustomerMatcher:
    def __init__(self, customers: List[CustomerRow]):
        self.customers = customers
        self._index: Dict[str, CustomerRow] = {}
        for c in customers:
            self._index[_norm(c.company_name)] = c

    def match(self, company_raw: str, threshold: float = 0.78) -> Tuple[Optional[CustomerRow], float]:
        key = _norm(company_raw)
        if key in self._index:
            return self._index[key], 1.0
        # fuzzy
        best, best_score = None, 0.0
        for c in self.customers:
            score = _ratio(key, _norm(c.company_name))
            if score > best_score:
                best, best_score = c, score
        if best_score >= threshold:
            return best, best_score
        return None, best_score


# ---------------------------------------------------------------------------
# PDF parsing
# ---------------------------------------------------------------------------

# The reports come in two layouts. We support both by extracting text and
# matching a flexible regex line-by-line.
#
# Layout A (TK / BS):
#   Company  Type  Gas  Amount  Vehicle  Station  Voucher  RedeemDate  Receipt
# Layout B (BL):
#   Company  Type  Amount  Gas  Litre  Vehicle  Station  Voucher  RedeemDate  Receipt
#
# We rely on:
#   - the amount being signed (always negative in source) and unique enough
#   - the redeem datetime in YYYY-MM-DD HH:MM:SS form
#   - the voucher matches \d{6}-\w+-\w+

VOUCHER_RX = re.compile(r"\d{6}-[A-Z0-9]+-[A-Z0-9]+")
DT_RX      = re.compile(r"20\d{2}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}")
AMOUNT_RX  = re.compile(r"-[\d,]+\.\d{2}")
GAS_RX     = re.compile(r"\b(Petrol|Diesel)\b", re.IGNORECASE)
STATION_RX = re.compile(r"Buraqoil\s+(Tg\s*Kapor|Berkat\s*Setia|Bubul\s*Lama)", re.IGNORECASE)

# v0.6.1: capture the per-PDF "Total:" footer so reconciliation can prove
# the parser didn't silently drop a redemption row. Matches "Total: -1,234.56"
# or "Grand Total: RM -1,234.56" etc. Sign and "RM" prefix are tolerated.
TOTAL_LINE_RX = re.compile(
    r"\b(?:Grand\s+)?Total\s*:?\s*(?:RM\s*)?-?\s*([\d,]+\.\d{2})",
    re.IGNORECASE,
)


@dataclass
class SourceParseResult:
    """Parser output enriched with the source's own control total.

    v0.6.1: introduced as PDFParseResult for the per-PDF Total: footer.
    v0.7.0: renamed -- now also covers the xlsx-source path, which carries
    the same control number as a 'Total: -36,848.03' string in the last
    data row of the sheet."""
    path: Path
    redemptions: List["Redemption"]
    source_total: Optional[float]  # value from the source's Total footer, abs-valued
    parsed_total: float            # sum of parsed Redemption.amount


def _extract_total(text: str) -> Optional[float]:
    """Return the numeric value from a 'Total:' line, or None if not a total line."""
    if "total" not in text.lower():
        return None
    m = TOTAL_LINE_RX.search(text)
    if not m:
        return None
    try:
        return float(m.group(1).replace(",", ""))
    except ValueError:
        return None


def parse_pdf(path: Path) -> SourceParseResult:
    """Extract redemptions AND the source 'Total:' line from a CFP/gvLedger PDF.

    The source total is captured from the LAST 'Total:' / 'Grand Total:' value
    seen anywhere in the document (rows, fallback text). If the PDF carries
    page subtotals, the bottom-of-document grand total wins -- which is what
    we want for the per-PDF reconciliation check A.
    """
    global pdfplumber
    if pdfplumber is None:
        try:
            import pdfplumber as _pp  # lazy import (v0.7.0)
            pdfplumber = _pp
        except ImportError:
            raise RuntimeError(
                "pdfplumber is required to parse PDF sources. "
                "Run: pip install pdfplumber  (or `pip install -r requirements.txt`). "
                "For xlsx-only sources, pdfplumber is not needed."
            )

    out: List[Redemption] = []
    source_total: Optional[float] = None

    with pdfplumber.open(str(path)) as pdf:
        for page in pdf.pages:
            # First try table extraction -- both layouts render as a
            # bordered table so this usually gets clean rows.
            tables = page.extract_tables() or []
            for table in tables:
                for row in table:
                    if not row:
                        continue
                    cells = [(c or "").strip() for c in row]
                    joined = " | ".join(cells)
                    t = _extract_total(joined)
                    if t is not None:
                        source_total = t  # last-seen wins
                        continue
                    rec = _row_to_redemption(row)
                    if rec:
                        out.append(rec)
            # Fallback: regex on raw text (handles malformed extractions)
            if not out:
                text = page.extract_text() or ""
                for line in text.splitlines():
                    t = _extract_total(line)
                    if t is not None:
                        source_total = t
                        continue
                    rec = _line_to_redemption(line)
                    if rec:
                        out.append(rec)

    parsed_total = round(sum(r.amount for r in out), 2)
    return SourceParseResult(path=path, redemptions=out,
                          source_total=source_total, parsed_total=parsed_total)


def _row_to_redemption(row: List[Optional[str]]) -> Optional[Redemption]:
    if not row:
        return None
    cells = [(c or "").strip() for c in row]
    joined = " | ".join(cells)
    if "Total:" in joined or "Company" in joined and "Voucher" in joined:
        return None
    m_dt   = DT_RX.search(joined)
    m_amt  = AMOUNT_RX.search(joined)
    m_vch  = VOUCHER_RX.search(joined)
    m_gas  = GAS_RX.search(joined)
    m_stn  = STATION_RX.search(joined)
    if not (m_dt and m_amt and m_vch and m_gas):
        return None
    company = cells[0]
    if not company or company.lower().startswith("company"):
        return None
    amount = float(m_amt.group().replace(",", "").replace("-", ""))
    # litre -- present in BL layout only; pick the first non-amount numeric
    litre = None
    for c in cells:
        if c and c != m_amt.group():
            mlit = re.fullmatch(r"-?\d{1,4}\.\d{2}", c)
            if mlit:
                v = float(c)
                if 0 < abs(v) < 5000:  # plausibly a litre count
                    litre = abs(v)
                    break
    vehicle = ""
    for c in cells:
        if re.match(r"^[A-Z]{1,3}\s*\d{1,5}\s*[A-Z/]*$", c) or re.match(r"^[A-Z]{2,3}\d{3,5}$", c):
            vehicle = c
            break
    station = "Buraqoil " + m_stn.group(1).title() if m_stn else ""
    return Redemption(
        company_raw=company,
        gas_type=m_gas.group(1).title(),
        amount=amount,
        litre=litre,
        vehicle=vehicle,
        station=station,
        voucher=m_vch.group(),
        redeem_dt=datetime.strptime(m_dt.group(), "%Y-%m-%d %H:%M:%S"),
        receipt=cells[-1],
    )


def _line_to_redemption(line: str) -> Optional[Redemption]:
    if "Total:" in line:
        return None
    m_dt   = DT_RX.search(line)
    m_amt  = AMOUNT_RX.search(line)
    m_vch  = VOUCHER_RX.search(line)
    m_gas  = GAS_RX.search(line)
    m_stn  = STATION_RX.search(line)
    if not (m_dt and m_amt and m_vch and m_gas):
        return None
    # Company is everything before the first "Refuel"
    head = re.split(r"\bRefuel\b", line, maxsplit=1)[0].strip()
    if not head:
        return None
    return Redemption(
        company_raw=head,
        gas_type=m_gas.group(1).title(),
        amount=float(m_amt.group().replace(",", "").replace("-", "")),
        litre=None,
        vehicle="",
        station="Buraqoil " + m_stn.group(1).title() if m_stn else "",
        voucher=m_vch.group(),
        redeem_dt=datetime.strptime(m_dt.group(), "%Y-%m-%d %H:%M:%S"),
        receipt="",
    )


# ---------------------------------------------------------------------------
# XLSX parsing  (v0.7.0)
# ---------------------------------------------------------------------------
#
# Unified gvLedger export: ONE sheet, ONE file covering all three stations.
# Sample shape (header row 1, data rows 2..N, Total row at N+1):
#
#   Company | Type | Amount | Gas | Litre | Vehicle | Station | Voucher | Redeem Date | Receipt
#   HIJAU MAJU JUTA SDN BHD | Refuel | -322.50 | Diesel | -150 | SYJ719 |
#       Buraqoil Berkat Setia | 260509-08-3V2IA | <serial> | 32247
#   ... 55 more redemption rows ...
#                       | Total: -36,848.03 |
#
# Sign convention observed in the source: Amount always negative (it's a
# debit from the voucher balance); Litre is mostly negative but Bubul Lama
# rows appear positive -- both are quantities, so we abs() them. Some rows
# carry Litre = 0 (system gap); those get derived via RefPrice in
# derive_missing_litres() with litre_estimated=True so the audit trail
# stays honest.

XLSX_COLUMN_ALIASES: Dict[str, List[str]] = {
    "company":  ["company"],
    "type":     ["type"],
    "amount":   ["amount"],
    "gas":      ["gas", "fuel", "product"],
    "litre":    ["litre", "liter", "litres", "liters"],
    "vehicle":  ["vehicle", "plate", "vehicleno"],
    "station":  ["station", "outlet"],
    "voucher":  ["voucher", "voucherno", "vouchercode"],
    "date":     ["redeemdate", "redeem date", "date", "datetime"],
    "receipt":  ["receipt", "receiptno"],
}


def _coerce_datetime(v) -> Optional[datetime]:
    """Best-effort: accept native datetime, Excel serial, or string."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, (int, float)):
        # Excel serial date (1900 system; openpyxl handles 1904 itself
        # when wb.epoch is set, but we read data_only and just shift).
        from datetime import timedelta
        return datetime(1899, 12, 30) + timedelta(days=float(v))
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S",
                    "%d/%m/%Y", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None


def _xlsx_normalize_station(s: str) -> str:
    """Source uses 'Buraqoil Tg Kapor' / 'Buraqoil Berkat Setia' /
    'Buraqoil Bubul Lama' verbatim -- matches STATION_NAME_TO_KEY keys."""
    return re.sub(r"\s+", " ", s).strip()


def parse_xlsx(path: Path) -> SourceParseResult:
    """v0.7.0: parse the unified gvLedger xlsx (one sheet, all stations).

    Returns SourceParseResult with redemptions and the source's Total
    footer for Check A. Rows with blank/zero Litre come out with
    `litre=None` -- derive_missing_litres() will fill them in after all
    sources are parsed (so we don't need RefPrices at parse time).
    """
    wb = load_workbook(str(path), data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Map header row -> column index (1-based)
    header_to_col: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        if v is not None:
            key = re.sub(r"\s+", " ", str(v).strip().lower())
            header_to_col[key] = col
            header_to_col[key.replace(" ", "")] = col  # also space-stripped

    def col_for(canonical: str) -> Optional[int]:
        for alias in XLSX_COLUMN_ALIASES[canonical]:
            if alias in header_to_col:
                return header_to_col[alias]
        return None

    c_company = col_for("company")
    c_amount  = col_for("amount")
    c_gas     = col_for("gas")
    c_litre   = col_for("litre")
    c_vehicle = col_for("vehicle")
    c_station = col_for("station")
    c_voucher = col_for("voucher")
    c_date    = col_for("date")
    c_receipt = col_for("receipt")

    missing = [k for k, c in [
        ("Company", c_company), ("Amount", c_amount), ("Gas", c_gas),
        ("Station", c_station), ("Voucher", c_voucher),
        ("Redeem Date", c_date),
    ] if c is None]
    if missing:
        raise ValueError(
            f"xlsx {path.name} is missing required column(s): {missing}. "
            f"Headers found: {sorted(header_to_col.keys())[:20]}..."
        )

    redemptions: List[Redemption] = []
    source_total: Optional[float] = None

    for row in ws.iter_rows(min_row=2, values_only=False):
        # Detect a Total row anywhere in the row (the sample has it as a
        # single 'Total: -36,848.03' string in the Amount cell). We also
        # capture standalone 'Total' / 'Grand Total' label + adjacent
        # number layouts.
        row_total_seen = False
        for cell in row:
            v = cell.value
            if v is None:
                continue
            s = str(v).strip()
            if not s:
                continue
            if "total" in s.lower():
                t = _extract_total(s)
                if t is not None:
                    source_total = t  # last-seen wins
                    row_total_seen = True
                    break
        if row_total_seen:
            continue

        def gv(c: Optional[int]):
            if c is None:
                return None
            return row[c - 1].value

        company = gv(c_company)
        if company is None or not str(company).strip():
            continue
        company = str(company).strip()

        try:
            amount = abs(float(gv(c_amount)))
        except (TypeError, ValueError):
            continue
        if amount <= 0:
            continue

        gas_raw = gv(c_gas)
        gas = str(gas_raw).strip().title() if gas_raw else ""
        if gas not in ("Petrol", "Diesel"):
            continue

        litre_raw = gv(c_litre)
        litre: Optional[float]
        try:
            litre_val = float(litre_raw) if litre_raw not in (None, "") else 0.0
        except (TypeError, ValueError):
            litre_val = 0.0
        litre = abs(litre_val) if litre_val != 0 else None  # 0 / missing -> derive later

        vehicle = str(gv(c_vehicle) or "").strip()
        station = _xlsx_normalize_station(str(gv(c_station) or ""))
        voucher = str(gv(c_voucher) or "").strip()
        receipt = str(gv(c_receipt) or "").strip()
        redeem_dt = _coerce_datetime(gv(c_date))
        if redeem_dt is None:
            continue

        redemptions.append(Redemption(
            company_raw=company,
            gas_type=gas,
            amount=amount,
            litre=litre,
            vehicle=vehicle,
            station=station,
            voucher=voucher,
            redeem_dt=redeem_dt,
            receipt=receipt,
        ))

    parsed_total = round(sum(r.amount for r in redemptions), 2)
    return SourceParseResult(path=path, redemptions=redemptions,
                             source_total=source_total, parsed_total=parsed_total)


def parse_source(path: Path) -> SourceParseResult:
    """v0.7.0: dispatch to PDF or xlsx parser by file extension.

    Both return the same SourceParseResult shape, so callers don't
    need to care which format they received."""
    ext = path.suffix.lower()
    if ext == ".pdf":
        return parse_pdf(path)
    if ext in (".xlsx", ".xlsm"):
        return parse_xlsx(path)
    raise ValueError(
        f"Unsupported source format: {ext} (path: {path}). "
        f"Supported: .pdf, .xlsx, .xlsm"
    )


def derive_missing_litres(redemptions: List[Redemption],
                          reference_prices: Dict[Tuple[str, str], float]) -> None:
    """v0.7.0: per-row Litre fallback using reference_prices.csv.

    Runs once after all sources are parsed. For every redemption with
    `litre is None or litre <= 0`, derives `litre = amount / RefPrice`
    and sets `litre_estimated = True`. RefPrice is looked up by
    (station_key, gas_type); if missing, the litre is left as None and
    the row will land in build_detail_rows with `est=True` but qty=0.

    This unifies three previously-separate cases:
      * TK / BS PDFs (entire source lacks a Litre column)
      * xlsx rows with a blank / zero Litre cell (per-row gap)
      * BL gvLedger PDFs (Litre column present; this function is a no-op)

    Behaviour for an all-litres-present bucket is unchanged. For a TK PDF
    bucket the per-row math is numerically identical to the v0.6.6
    bucket-level fallback (sum of (amount_i / RefPrice) == sum(amount) /
    RefPrice). For a mixed xlsx bucket -- some rows with litres, some
    without -- this is the first version that handles them correctly.
    Mutates `redemptions` in place.
    """
    for r in redemptions:
        if r.litre is not None and r.litre > 0:
            continue
        stn_key = STATION_NAME_TO_KEY.get(r.station, "")
        rp = reference_prices.get((stn_key, r.gas_type), 0.0)
        if rp > 0:
            r.litre = r.amount / rp
            r.litre_estimated = True
        # else: leave litre=None; downstream will mark qty=0 and est=True


# ---------------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------------

@dataclass
class ConsolidatedRow:
    doc_date: str
    acc_code: str
    company_name: str
    station_key: str      # TK / BS / BL  (v0.2.0 -- now part of group key)
    gas_type: str         # Petrol / Diesel
    total_amount: float   # positive
    total_litre: float    # 0.0 if all source rows had no litre info
    voucher_count: int
    voucher_list: str     # ;-separated for audit trail
    station_full: str     # human-readable e.g. "Buraqoil Tg Kapor"
    notes: str = ""
    # v0.6.3: source rows that contributed to this bucket -- needed for
    # per-row unit-price reconciliation (Check C).
    source_redemptions: List[Redemption] = field(default_factory=list)


def _station_key_from_full(s: str) -> str:
    return STATION_NAME_TO_KEY.get(s, "")


def consolidate(redemptions: Iterable[Redemption],
                matcher: CustomerMatcher) -> Tuple[List[ConsolidatedRow], List[Redemption]]:
    """v0.2.0: group key = (date, acc_code, station, gas_type)."""
    bucket: Dict[Tuple[str, str, str, str], ConsolidatedRow] = {}
    unmapped: List[Redemption] = []
    for r in redemptions:
        cust, score = matcher.match(r.company_raw)
        if cust is None:
            unmapped.append(r)
            continue
        stn_key = _station_key_from_full(r.station)
        key = (r.date, cust.acc_code, stn_key, r.gas_type)
        if key not in bucket:
            bucket[key] = ConsolidatedRow(
                doc_date=r.date,
                acc_code=cust.acc_code,
                company_name=cust.company_name,
                station_key=stn_key,
                gas_type=r.gas_type,
                total_amount=0.0,
                total_litre=0.0,
                voucher_count=0,
                voucher_list="",
                station_full=r.station,
                notes="" if score >= 0.99 else f"fuzzy match {score:.2f} from '{r.company_raw}'",
            )
        cr = bucket[key]
        cr.total_amount += r.amount
        cr.total_litre  += (r.litre or 0.0)
        cr.voucher_count += 1
        cr.voucher_list = f"{cr.voucher_list};{r.voucher}".strip(";")
        cr.source_redemptions.append(r)  # v0.6.3: track for per-row recon
    rows = sorted(bucket.values(),
                  key=lambda x: (x.doc_date, x.acc_code, x.station_key, x.gas_type))
    return rows, unmapped


# ---------------------------------------------------------------------------
# AutoCount xlsx writer
# ---------------------------------------------------------------------------

# AutoCount's "Default Sales.xlsx" template has Master and Detail sheets.
# Rather than hard-code them, we open the template, read row-1 headers
# verbatim, and only fill the columns that actually exist. Anything we
# don't have data for is left blank. The user can extend by editing the
# COLUMN_MAP below.

# Lower-cased canonical name -> list of header keywords any of which counts.
# Add aliases as your AutoCount build differs.
COLUMN_MAP_MASTER = {
    "doc_no":       ["docno", "doc no", "invoice no", "invoiceno"],
    "doc_date":     ["docdate", "doc date", "invoice date", "date"],
    "debtor_code":  ["debtorcode", "debtor code", "account no", "account code", "code"],
    "debtor_name":  ["debtorname", "debtor name", "customer name", "name"],
    "description":  ["description", "remark", "memo", "note"],
    "agent":        ["agent"],
    "currency":     ["currencycode", "currency"],
    "exchange":     ["exchangerate", "rate"],
}
COLUMN_MAP_DETAIL = {
    "doc_no":              ["docno", "doc no", "invoice no"],
    "item_code":           ["itemcode", "item code", "stock code", "stockcode"],
    "description":         ["description", "remark"],
    # v0.6.5: AutoCount Detail templates carry a separate
    # FurtherDescription field (longer audit trail). The voucher list
    # rides here, so the audit trail lands inside AutoCount on import
    # rather than living only in the side Audit sheet.
    "further_description": ["furtherdescription", "furtherdesc",
                            "further description", "further desc"],
    "uom":                 ["uom"],
    "qty":                 ["qty", "quantity", "numofunit"],
    "unit_price":          ["unitprice", "price"],
    "discount":            ["discount"],
    "subtotal":            ["subtotal", "amount", "total"],
    "tax_code":            ["taxcode", "taxtype"],
    "tax_amount":          ["taxamount", "tax amount"],
    "project_no":          ["projectno", "project no", "project"],
}


def _find_headers(ws, alias_map: Dict[str, List[str]]) -> Dict[str, int]:
    """Return canonical_name -> column_index (1-based) by scanning row 1.

    v0.6.5: two-pass matching to avoid `description` accidentally
    matching `FurtherDescription` (substring `"description" in
    "furtherdescription"` is True).

      Pass 1 -- EXACT match against either the lowercased header or
                the space-stripped lowercased header. A column claimed
                here is removed from the pool for Pass 2.
      Pass 2 -- SUBSTRING fallback for canonicals that didn't find an
                exact match. Still respects the claim set.

    With this, when both `Description` and `FurtherDescription` exist
    in the template, `description` resolves to the exact `Description`
    column (Pass 1) and `further_description` resolves to
    `FurtherDescription` (also Pass 1) -- no cross-talk.
    """
    found: Dict[str, int] = {}
    if ws.max_row < 1:
        return found

    headers: List[Tuple[int, str, str]] = []  # (col, lowered, no-space)
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=col).value
        h = str(v).strip().lower() if v else ""
        headers.append((col, h, h.replace(" ", "")))

    claimed: set = set()

    # Pass 1: exact match (preferred)
    for canonical, aliases in alias_map.items():
        for col, h, h_ns in headers:
            if not h or col in claimed:
                continue
            if any(a == h or a == h_ns for a in aliases):
                found[canonical] = col
                claimed.add(col)
                break

    # Pass 2: substring fallback for unmapped canonicals only
    for canonical, aliases in alias_map.items():
        if canonical in found:
            continue
        for col, h, h_ns in headers:
            if not h or col in claimed:
                continue
            if any(a in h or a in h_ns for a in aliases):
                found[canonical] = col
                claimed.add(col)
                break

    return found


@dataclass
class DetailRow:
    """v0.6.1: precomputed detail row -- the SAME numbers used for both
    write_import (xlsx output) and build_reconciliation (verification).
    Decoupling these guarantees the reconciliation reads what was actually
    written, not a re-derivation that could drift.

    v0.6.5: further_description added for the AutoCount FurtherDescription
    column. Holds the voucher list so the audit trail rides into AutoCount
    on import."""
    doc_no: str
    cr: "ConsolidatedRow"
    item_code: str
    project_no: str
    description: str
    further_description: str
    uom: str
    qty: float
    unit_price: float
    subtotal: float
    # v0.7.0: True iff EVERY source row in the bucket was estimated (full
    # RefPrice fallback). For a mixed bucket (some source litres, some
    # derived) this is False; use `est_count` on the DetailRow or check
    # any(r.litre_estimated for r in cr.source_redemptions) for finer
    # granularity. Kept for backwards-compat with anything that read
    # this flag from the Audit sheet / programmatic callers.
    qty_estimated: bool
    # v0.7.0: count of source rows in this bucket that had Litre derived
    # via RefPrice. 0 = clean source-truth; equal to voucher_count =
    # full estimation; in between = mixed (xlsx-only case).
    est_count: int = 0


def _format_voucher_line(r: Redemption) -> str:
    """v0.6.6+: one voucher line for the FurtherDescription audit trail.

    Format: '<voucher> - <fuel> (<qty>)' where <qty> is verbatim from
    the source -- NEVER derived:
      * Source had a real litre (BL gvLedger row, xlsx row with Litre>0)
        -> '25.19L'
      * Source had no litre or a zero litre, so the value was estimated
        via RefPrice in derive_missing_litres()
        -> 'RM 100.00'    (per-row amount, NOT the derived litre)

    v0.7.0 honesty rule: after derive_missing_litres() runs, every
    redemption has a numeric `litre`. The old check (`r.litre is not None
    and r.litre > 0`) would silently print a derived value as if it were
    source-truth. The new check uses `litre_estimated` so estimated rows
    keep showing the original RM amount -- aligning with the v0.6.6 docs
    promise that <qty> is verbatim from the source.
    """
    if r.litre is None or r.litre <= 0 or r.litre_estimated:
        qty_text = f"RM {r.amount:.2f}"
    else:
        qty_text = f"{r.litre:.2f}L"
    return f"{r.voucher} - {r.gas_type} ({qty_text})"


def _build_further_description(cr: ConsolidatedRow, cap: int = 500) -> str:
    """v0.6.6: build the date-grouped voucher trail for FurtherDescription.

    Format:
        YYYY-MM-DD:
        <voucher> - <fuel> (<qty>)
        <voucher> - <fuel> (<qty>)
        ...

        YYYY-MM-DD:
        <voucher> - <fuel> (<qty>)
        ...

    Sections separated by a blank line; vouchers within each date sorted
    by redemption datetime so the trail reads chronologically.

    Under the v0.3.0 consolidation rule (one bucket per
    date+customer+station+fuel), every bucket has exactly one date and
    one fuel, so the rendered text has a single date heading. The
    date-grouping logic remains so that if consolidation is ever
    relaxed (e.g., (customer, station)-only), the same code adapts
    without modification.

    Capped at 500 chars (typical AutoCount FurtherDescription column
    width); anything longer is truncated with a '...' suffix.
    """
    by_date: Dict[str, List[Redemption]] = {}
    for r in cr.source_redemptions:
        by_date.setdefault(r.date, []).append(r)

    sections: List[str] = []
    for date in sorted(by_date.keys()):
        rows = sorted(by_date[date], key=lambda r: r.redeem_dt)
        lines = [f"{date}:"]
        for r in rows:
            lines.append(_format_voucher_line(r))
        sections.append("\n".join(lines))

    text = "\n\n".join(sections)
    if len(text) > cap:
        text = text[: cap - 3] + "..."
    return text


def build_detail_rows(rows: List[ConsolidatedRow],
                      stock_codes: Dict[Tuple[str, str], str],
                      project_codes: Dict[str, str],
                      reference_prices: Dict[Tuple[str, str], float],
                      doc_no_prefix: str = "CFP") -> List[DetailRow]:
    """Compute the AutoCount detail row for every consolidated row.

    Header and detail are 1-to-1 in v0.3.0+, so the DocNo emitted here is
    the same DocNo the master row will carry in write_import.
    """
    out: List[DetailRow] = []
    seq = 1
    for cr in rows:
        doc_no = f"{doc_no_prefix}-{cr.doc_date.replace('-', '')}-{seq:04d}"
        seq += 1
        item_code = stock_codes.get((cr.station_key, cr.gas_type), "")
        project_no = project_codes.get(cr.station_key, "")

        # v0.6.4: UnitPrice is taken DIRECTLY from a source row's
        # amount/litre (rounded 2 dp). NEVER a weighted average. The CFP
        # system is automatically generated so per-row amount/litre in a
        # single (date, customer, station, fuel) bucket is guaranteed to
        # match. Reconciliation Check C enforces this and FAILs the run
        # if any row disagrees -- a disagreement means corrupt source
        # data, not a legitimate price change.
        #
        # v0.7.0: per-row Litre fallback. After derive_missing_litres()
        # runs, every redemption has a numeric `litre`. The bucket may
        # now be:
        #   * all-source  -- every row's litre came from the source
        #                    (e.g. xlsx file with no Litre=0 rows, or a
        #                    BL gvLedger PDF). UnitPrice = first source
        #                    row's amount/litre.
        #   * all-estimated -- every row had litre derived from RefPrice
        #                    (TK / BS PDFs, or all-blank xlsx bucket).
        #                    UnitPrice = round(RefPrice, 2) directly.
        #   * mixed       -- xlsx bucket where some rows have litres and
        #                    some don't. UnitPrice = first SOURCE row's
        #                    amount/litre; Qty = sum of all litres
        #                    (source + derived). Description gets a
        #                    "(N vch, K est)" suffix to flag the mix.
        # `est_count` is the number of estimated rows in the bucket;
        # 0 = clean, == n = full estimation, in between = mixed.
        source_priced_rows = [r for r in cr.source_redemptions
                              if r.unit_price_per_row is not None]
        est_count = sum(1 for r in cr.source_redemptions if r.litre_estimated)
        # Qty: sum across all rows (source litres + derived litres).
        # cr.total_litre is already this sum because consolidate() adds
        # r.litre on every row, and derive_missing_litres() filled in
        # the gaps before consolidate ran.
        if cr.total_litre and cr.total_litre > 0:
            qty = round(cr.total_litre, 2)
        else:
            qty = 0.0
        # UnitPrice: prefer a source row's per-row price (preserves the
        # v0.6.4 "never weighted average" rule). If the bucket is fully
        # estimated, fall back to the RefPrice itself.
        if source_priced_rows:
            unit_price = source_priced_rows[0].unit_price_per_row or 0.0
        else:
            ref_price = reference_prices.get((cr.station_key, cr.gas_type), 0.0)
            unit_price = round(ref_price, 2) if ref_price > 0 else 0.0
        subtotal = round(cr.total_amount, 2)

        # v0.6.0 / v0.7.0: short description, no voucher list (80-char cap).
        #   no estimation         -> "Petrol @ ... (5 vch)"
        #   fully estimated       -> "Petrol @ ... (5 vch est)"        [v0.6.0 form]
        #   mixed estimation      -> "Petrol @ ... (5 vch, 2 est)"     [v0.7.0]
        n = cr.voucher_count
        if est_count == 0:
            est_suffix = ""
        elif est_count == n:
            est_suffix = " est"
        else:
            est_suffix = f", {est_count} est"
        desc = (f"{cr.gas_type} @ {cr.station_full or cr.station_key} "
                f"({n} vch{est_suffix})")[:80]

        # v0.6.6: FurtherDescription is now date-grouped voucher detail.
        # See _build_further_description() for format spec; cap raised
        # to 500 chars since the new format is more verbose.
        further_desc = _build_further_description(cr)

        out.append(DetailRow(
            doc_no=doc_no,
            cr=cr,
            item_code=item_code,
            project_no=project_no,
            description=desc,
            further_description=further_desc,
            uom="LITER",
            qty=qty,
            unit_price=unit_price,
            subtotal=subtotal,
            qty_estimated=(est_count == n),  # all-or-nothing flag (v0.6.0 semantic)
            est_count=est_count,             # v0.7.0: granular count
        ))
    return out


def write_import(template_path: Path,
                 out_path: Path,
                 rows: List[ConsolidatedRow],
                 detail_rows: List[DetailRow],
                 unmapped: List[Redemption],
                 stock_codes: Dict[Tuple[str, str], str],
                 project_codes: Dict[str, str],
                 reconciliation: Optional[List["ReconCheck"]] = None) -> None:
    """v0.6.1: takes pre-built DetailRow objects (so reconciliation sees the
    exact same numbers) and an optional reconciliation report (written to a
    `Reconciliation` sheet for the auditor)."""
    wb = load_workbook(str(template_path))
    sheet_names = wb.sheetnames

    # Locate master & detail sheets
    master_ws = wb[sheet_names[0]]
    detail_ws = wb[sheet_names[1]] if len(sheet_names) > 1 else None

    master_cols = _find_headers(master_ws, COLUMN_MAP_MASTER)
    detail_cols = _find_headers(detail_ws, COLUMN_MAP_DETAIL) if detail_ws else {}

    # Clear any existing sample data below row 1
    for ws in (master_ws, detail_ws):
        if ws is None:
            continue
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.value = None

    # v0.3.0: ONE Sales Invoice per (date, debtor, station, fuel).
    # Header and detail are 1-to-1.
    next_master_row = 2
    next_detail_row = 2
    for d in detail_rows:
        cr = d.cr
        # Master row
        mr = next_master_row
        next_master_row += 1
        master_desc = (f"CFP {cr.gas_type} @ {cr.station_full or cr.station_key} "
                       f"{cr.doc_date}")
        _set(master_ws, mr, master_cols.get("doc_no"),       d.doc_no)
        _set(master_ws, mr, master_cols.get("doc_date"),     cr.doc_date)
        _set(master_ws, mr, master_cols.get("debtor_code"),  cr.acc_code)
        _set(master_ws, mr, master_cols.get("debtor_name"),  cr.company_name)
        _set(master_ws, mr, master_cols.get("description"),  master_desc)
        _set(master_ws, mr, master_cols.get("currency"),     "RM")
        _set(master_ws, mr, master_cols.get("exchange"),     1)

        # Detail row
        if detail_ws is None:
            continue
        dr = next_detail_row
        next_detail_row += 1
        _set(detail_ws, dr, detail_cols.get("doc_no"),              d.doc_no)
        _set(detail_ws, dr, detail_cols.get("item_code"),           d.item_code)
        _set(detail_ws, dr, detail_cols.get("description"),         d.description)
        _set(detail_ws, dr, detail_cols.get("further_description"), d.further_description)
        _set(detail_ws, dr, detail_cols.get("uom"),                 d.uom)
        _set(detail_ws, dr, detail_cols.get("qty"),                 d.qty)
        _set(detail_ws, dr, detail_cols.get("unit_price"),          d.unit_price)
        _set(detail_ws, dr, detail_cols.get("subtotal"),            d.subtotal)
        if d.project_no:
            _set(detail_ws, dr, detail_cols.get("project_no"), d.project_no)

    # Audit sheet (always)
    if "Audit" in wb.sheetnames:
        audit_ws = wb["Audit"]
        wb.remove(audit_ws)
    audit_ws = wb.create_sheet("Audit")
    audit_ws.append(["Date", "AccCode", "Customer", "Station", "Gas",
                     "ItemCode", "ProjectNo", "Amount", "Litre",
                     "VoucherCount", "EstCount", "Vouchers", "Notes"])
    for cr in rows:
        # v0.7.0: EstCount surfaces per-row litre estimation in the audit
        # trail. 0 = bucket is clean source-truth. == VoucherCount = full
        # RefPrice fallback (TK / BS PDF, or all-blank xlsx bucket).
        # In between = mixed xlsx bucket (some source litres, some derived).
        est_count = sum(1 for r in cr.source_redemptions if r.litre_estimated)
        audit_ws.append([cr.doc_date, cr.acc_code, cr.company_name,
                         cr.station_full or cr.station_key, cr.gas_type,
                         stock_codes.get((cr.station_key, cr.gas_type), ""),
                         project_codes.get(cr.station_key, ""),
                         round(cr.total_amount, 2), round(cr.total_litre, 2),
                         cr.voucher_count, est_count,
                         cr.voucher_list, cr.notes])

    # Unmapped sheet
    if "Unmapped" in wb.sheetnames:
        wb.remove(wb["Unmapped"])
    unmapped_ws = wb.create_sheet("Unmapped")
    unmapped_ws.append(["CompanyAsReported", "Gas", "Amount", "Voucher",
                        "Station", "RedeemDate", "Receipt"])
    for r in unmapped:
        unmapped_ws.append([r.company_raw, r.gas_type, r.amount, r.voucher,
                            r.station, r.redeem_dt.strftime("%Y-%m-%d %H:%M:%S"),
                            r.receipt])

    # v0.6.1: Reconciliation sheet -- written even if all checks pass, so
    # the operator (or auditor) always has a paper trail of the totals.
    if reconciliation is not None:
        if "Reconciliation" in wb.sheetnames:
            wb.remove(wb["Reconciliation"])
        rec_ws = wb.create_sheet("Reconciliation")
        rec_ws.append(["Status", "Check", "Expected", "Actual", "Delta", "Note"])
        for c in reconciliation:
            rec_ws.append([
                c.status,
                c.name,
                "" if c.expected is None else round(c.expected, 2),
                "" if c.actual is None else round(c.actual, 2),
                "" if c.delta is None else round(c.delta, 2),
                c.note,
            ])

    # v0.6.2: AutoCount computes its own totals on import. The output
    # MUST NOT carry a Total row in either Master or Detail. Scan the
    # rows we just wrote (skip row 1 = headers) and abort if one slipped
    # through, e.g. a stale Total row in the template that survived the
    # row-2-onwards clear.
    _assert_no_total_row(master_ws, "Master")
    if detail_ws is not None:
        _assert_no_total_row(detail_ws, "Detail")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


def _assert_no_total_row(ws, sheet_label: str) -> None:
    """Raise if any data row (row >= 2) starts with 'Total' or contains
    a cell whose value is exactly 'Total' / 'Grand Total' (case-insensitive).
    AutoCount imports must contain only header + transaction rows."""
    bad: List[int] = []
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            s = str(v).strip().lower().rstrip(":")
            if s in {"total", "grand total", "sub total", "subtotal"}:
                bad.append(cell.row)
                break
    if bad:
        raise RuntimeError(
            f"{sheet_label} sheet contains Total row(s) at row(s) {bad}. "
            f"AutoCount computes its own totals on import -- output must "
            f"contain only header + transaction rows. Aborting save."
        )


def _set(ws, row: int, col: Optional[int], value):
    if col is None or row is None:
        return
    ws.cell(row=row, column=col, value=value)


# ---------------------------------------------------------------------------
# Reconciliation  (v0.6.1)
# ---------------------------------------------------------------------------
#
# Every CFP run must prove three things before the xlsx is safe to import:
#
#   A. PARSE INTEGRITY (per-PDF)
#      The PDF's own "Total:" footer line equals the sum of redemption
#      rows the parser extracted. If the parser drops a row (regex miss,
#      odd cell layout) this check goes red. Skipped only when the PDF
#      doesn't carry a Total line at all.
#
#   B. PIPELINE CONSERVATION
#      Sum of filtered redemptions (after --date filter) must equal
#      sum of consolidated rows + sum of unmapped rows. This proves the
#      consolidator and customer-matcher together didn't lose money.
#
#   C. PER-ROW UNIT-PRICE UNIFORMITY  (v0.6.3+, hardened in v0.6.4)
#      For every source redemption with a litre value, the bucket's
#      UnitPrice (taken directly from one source row's amount/litre,
#      2 dp) must equal that row's amount/litre within RM 0.01.
#      The CFP system is automatically generated, so per-row prices
#      in one bucket are deterministically uniform. A disagreement
#      means corrupt source data -- never a legitimate mid-day price
#      change -- and the run MUST fail. The RefPrice fallback path
#      is exempt because no per-row litre data is available.
#
#      v0.6.4 explicitly forbids weighted-average UnitPrice. If a
#      bucket contains rows with disagreeing per-row prices, the
#      bucket UnitPrice is the FIRST row's value (deterministic),
#      and Check C surfaces every other row as a FAIL.
#
#      The old "Qty * UnitPrice ~ Subtotal" check was meaningful when
#      UnitPrice was 4 dp; at 2 dp the rounding error scales with Qty
#      and the check would routinely report ~0.40 RM drift on normal
#      buckets. AutoCount uses the explicit Subtotal column on import,
#      and pipeline conservation (Check B) already guarantees
#      Sum(detail Subtotal) == Sum(filtered amount) - Sum(unmapped).
#
# Tolerance is RM 0.01 across the board (cent-level). Any FAIL causes the
# script to exit with code 2 -- the operator must NOT import the xlsx
# until the failure is understood.
# ---------------------------------------------------------------------------

RECON_TOLERANCE = 0.01


@dataclass
class ReconCheck:
    name: str
    expected: Optional[float]
    actual: Optional[float]
    delta: Optional[float]
    status: str        # "OK" / "FAIL" / "SKIP"
    note: str = ""


def _check(name: str, expected: float, actual: float,
           tol: float = RECON_TOLERANCE, note: str = "") -> ReconCheck:
    delta = round(actual - expected, 4)
    status = "OK" if abs(delta) < tol else "FAIL"
    return ReconCheck(name=name, expected=expected, actual=actual,
                      delta=delta, status=status, note=note)


def build_reconciliation(source_results: List[SourceParseResult],
                         filtered: List[Redemption],
                         consolidated: List[ConsolidatedRow],
                         unmapped: List[Redemption],
                         detail_rows: List[DetailRow]) -> List[ReconCheck]:
    checks: List[ReconCheck] = []

    # --- Check A: per-source parse integrity -------------------------------
    for pr in source_results:
        label = f"A. {pr.path.name}: source Total vs parsed sum"
        if pr.source_total is None:
            checks.append(ReconCheck(
                name=label, expected=None, actual=pr.parsed_total,
                delta=None, status="SKIP",
                note="no 'Total:' line found in PDF; parser sum reported as actual",
            ))
        else:
            checks.append(_check(label, pr.source_total, pr.parsed_total))

    # --- Check B: pipeline conservation ------------------------------------
    filtered_total     = round(sum(r.amount for r in filtered),            2)
    consolidated_total = round(sum(cr.total_amount for cr in consolidated), 2)
    unmapped_total     = round(sum(r.amount for r in unmapped),             2)
    checks.append(_check(
        name="B. Conservation: filtered = consolidated + unmapped",
        expected=filtered_total,
        actual=round(consolidated_total + unmapped_total, 2),
        note=f"consolidated {consolidated_total:.2f} + unmapped {unmapped_total:.2f}",
    ))

    # --- Check C: per-row UnitPrice uniformity (v0.6.3, v0.7.0 updated) ----
    # For every source row with a REAL (non-estimated) litre value, the
    # bucket's UnitPrice must equal round(amount / litre, 2) within
    # RM 0.01. Rows where the litre was derived via RefPrice in
    # derive_missing_litres() are exempt -- their unit_price_per_row
    # property returns None, so they're skipped naturally.
    #
    # v0.7.0 nuance: a mixed bucket (some source litres, some estimated)
    # is checked only on the source-litre rows. The estimated rows are
    # counted in the OK note for transparency, but they can't disagree
    # by construction (amount / (amount/RefPrice) = RefPrice).
    bad: List[Tuple[str, str, float, float]] = []
    rows_checked = 0
    rows_skipped_estimated = 0
    for d in detail_rows:
        for r in d.cr.source_redemptions:
            row_up = r.unit_price_per_row
            if row_up is None:
                if r.litre_estimated:
                    rows_skipped_estimated += 1
                continue
            rows_checked += 1
            if abs(row_up - d.unit_price) > RECON_TOLERANCE:
                bad.append((d.doc_no, r.voucher, row_up, d.unit_price))
    if bad:
        sample = "; ".join(
            f"{doc} {vch}: row {ru:.2f} vs bucket {bu:.2f}"
            for doc, vch, ru, bu in bad[:3]
        )
        more = "" if len(bad) <= 3 else f" (+{len(bad) - 3} more)"
        checks.append(ReconCheck(
            name="C. Per-row UnitPrice uniformity: amount/litre matches bucket UnitPrice",
            expected=0.0, actual=float(len(bad)), delta=float(len(bad)),
            status="FAIL",
            note=f"{len(bad)} row(s) outside RM {RECON_TOLERANCE:.2f}: {sample}{more}",
        ))
    else:
        note_bits = [f"{rows_checked} source row(s) checked"]
        if rows_skipped_estimated:
            note_bits.append(f"{rows_skipped_estimated} estimated row(s) skipped")
        checks.append(ReconCheck(
            name="C. Per-row UnitPrice uniformity: amount/litre matches bucket UnitPrice",
            expected=0.0, actual=0.0, delta=0.0, status="OK",
            note="; ".join(note_bits),
        ))

    return checks


def print_reconciliation(checks: List[ReconCheck]) -> None:
    print("\n=== Reconciliation ============================================")
    for c in checks:
        exp = f"{c.expected:>12.2f}" if c.expected is not None else "          --"
        act = f"{c.actual:>12.2f}"   if c.actual   is not None else "          --"
        dlt = f"{c.delta:+10.2f}"    if c.delta    is not None else "        --"
        print(f"  [{c.status:4s}] {c.name}")
        print(f"          expected={exp}  actual={act}  delta={dlt}")
        if c.note:
            print(f"          note: {c.note}")
    fails = [c for c in checks if c.status == "FAIL"]
    print(f"\n  {len(fails)} FAIL, "
          f"{sum(1 for c in checks if c.status == 'OK')} OK, "
          f"{sum(1 for c in checks if c.status == 'SKIP')} SKIP")
    if fails:
        print("  >>> DO NOT IMPORT into AutoCount until every FAIL is resolved.")
    print("================================================================\n")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(description="Compile CFP voucher redemptions into an AutoCount sales import xlsx.")
    p.add_argument("--reports", nargs="+", required=True,
                   help="One or more CFP/gvLedger source files (.pdf or .xlsx). "
                        "v0.7.0+: a single unified gvLedger.xlsx covering all "
                        "three stations is the recommended format.")
    p.add_argument("--date", default=None, help="Document date YYYY-MM-DD (auto-detect if omitted)")
    p.add_argument("--template", required=True, help="Path to AutoCount Default Sales.xlsx")
    p.add_argument("--customers", required=True, help="customer_codes.csv")
    p.add_argument("--out", required=True, help="Output xlsx path")
    p.add_argument("--stock-codes", default=None,
                   help="Path to stock_codes.csv (overrides built-in mapping)")
    p.add_argument("--project-codes", default=None,
                   help="Path to project_codes.csv")
    p.add_argument("--reference-prices", default=None,
                   help="Path to reference_prices.csv (per-station-gas RM/L)")
    # deprecated -- silently accepted for backward compat
    p.add_argument("--petrol-code", default=None, help=argparse.SUPPRESS)
    p.add_argument("--diesel-code", default=None, help=argparse.SUPPRESS)
    p.add_argument("--doc-prefix", default="CFP", help="DocNo prefix (default: CFP)")
    args = p.parse_args(argv)

    if args.petrol_code or args.diesel_code:
        print("WARNING: --petrol-code / --diesel-code are deprecated since v0.2.0; "
              "the script picks per-station codes from stock_codes.csv.",
              file=sys.stderr)

    matcher = CustomerMatcher(load_customers(Path(args.customers)))
    stock_codes = load_stock_codes(Path(args.stock_codes) if args.stock_codes else None)
    project_codes = load_project_codes(Path(args.project_codes) if args.project_codes else None)
    reference_prices = load_reference_prices(
        Path(args.reference_prices) if args.reference_prices else None)

    source_results: List[SourceParseResult] = []
    all_red: List[Redemption] = []
    for rp in args.reports:
        path = Path(rp)
        if not path.exists():
            print(f"WARN: source not found: {rp}", file=sys.stderr)
            continue
        try:
            pr = parse_source(path)
        except ValueError as e:
            print(f"WARN: {e}", file=sys.stderr)
            continue
        source_results.append(pr)
        all_red.extend(pr.redemptions)
        if pr.source_total is None:
            print(f"  parsed {len(pr.redemptions):3d} rows  parsed_total={pr.parsed_total:>12.2f}  "
                  f"source_total=  (no Total: line)   <- {path.name}")
        else:
            delta = round(pr.parsed_total - pr.source_total, 2)
            mark = "OK" if abs(delta) < RECON_TOLERANCE else "MISMATCH"
            print(f"  parsed {len(pr.redemptions):3d} rows  parsed_total={pr.parsed_total:>12.2f}  "
                  f"source_total={pr.source_total:>12.2f}  delta={delta:+.2f} [{mark}]   <- {path.name}")

    if args.date:
        all_red = [r for r in all_red if r.date == args.date]
        print(f"Filtered to {args.date}: {len(all_red)} rows")

    # v0.7.0: post-parse Litre normalisation. Rows that came out of the
    # parser without a usable litre (TK / BS PDF rows, xlsx rows with a
    # blank / zero Litre cell) get litre derived from RefPrice and
    # marked litre_estimated=True. Downstream code (consolidate,
    # build_detail_rows, reconciliation) is then one path -- every row
    # has a numeric litre, the estimated flag drives the cosmetic
    # differences (description suffix, voucher line format,
    # Check C exemption).
    estimated_before = sum(1 for r in all_red if r.litre is None or r.litre <= 0)
    derive_missing_litres(all_red, reference_prices)
    if estimated_before:
        print(f"Estimated litre via RefPrice for {estimated_before} row(s) "
              f"(no source litre)")

    rows, unmapped = consolidate(all_red, matcher)
    print(f"\nConsolidated {len(rows)} customer-fuel rows; {len(unmapped)} unmapped rows.")
    if unmapped:
        print("Unmapped companies (add to customer_codes.csv):")
        for r in unmapped:
            print(f"   - {r.company_raw}  ({r.gas_type}, {r.amount:.2f})")

    # v0.6.1: precompute detail rows so reconciliation reads the SAME
    # numbers that get written to xlsx.
    detail_rows = build_detail_rows(rows, stock_codes, project_codes,
                                    reference_prices, args.doc_prefix)

    # v0.6.1: build + print reconciliation BEFORE writing the xlsx so the
    # operator sees totals even if the xlsx write fails for any reason.
    recon = build_reconciliation(source_results, all_red, rows, unmapped, detail_rows)
    print_reconciliation(recon)

    write_import(
        template_path=Path(args.template),
        out_path=Path(args.out),
        rows=rows,
        detail_rows=detail_rows,
        unmapped=unmapped,
        stock_codes=stock_codes,
        project_codes=project_codes,
        reconciliation=recon,
    )
    print(f"Wrote import file: {args.out}")

    # v0.6.1: hard fail on any reconciliation FAIL so an automated caller
    # (CI / scheduler / Claude) can't silently import a broken xlsx.
    fails = [c for c in recon if c.status == "FAIL"]
    if fails:
        print(f"\nERROR: {len(fails)} reconciliation check(s) failed. "
              f"DO NOT IMPORT the xlsx until resolved.", file=sys.stderr)
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
